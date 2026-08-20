"""Optional spreadsheet parsers with sheet, row, and column provenance."""

from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import ClassVar

from ...models import DocTree, Document, SourceLocation
from ..builder import DocTreeBuilder


def _title(document: Document) -> str:
    if document.document_title:
        return document.document_title
    if document.source_path is not None:
        return Path(document.source_path).stem or "Anonymous"
    return "Anonymous"


def _source_path(document: Document) -> str | None:
    return str(document.source_path) if document.source_path is not None else None


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    return str(value)


class XlsxParser:
    """Parse each non-empty XLSX row as an atomic, sheet-aware record.

    Requires ``pip install lumberjack-py[spreadsheets]``.
    """

    default_block_kinds: ClassVar[frozenset[str]] = frozenset({"tabular_row"})

    @property
    def block_kinds(self) -> frozenset[str]:
        return self.default_block_kinds

    def parse(
        self,
        document: Document | bytes,
        *,
        document_title: str | None = None,
        metadata_overrides: dict[str, object] | None = None,
        source_path: str | Path | None = None,
    ) -> DocTree:
        if not isinstance(document, Document):
            document = Document(
                document,
                format="xlsx",
                document_title=document_title,
                metadata_overrides=dict(metadata_overrides or {}),
                source_path=source_path,
            )
        if not isinstance(document.source, bytes | bytearray):
            raise TypeError("XlsxParser.parse expects Document[bytes]")
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "XLSX support requires 'openpyxl'. Install lumberjack-py[spreadsheets]."
            ) from exc

        workbook = load_workbook(
            BytesIO(bytes(document.source)), read_only=True, data_only=False
        )
        builder = DocTreeBuilder(
            title=_title(document),
            source="",
            source_path=_source_path(document),
            metadata=document.metadata_overrides,
            block_kinds=self.block_kinds,
            topology="records",
        )
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            try:
                header_values = next(rows)
            except StopIteration:
                continue
            header = tuple(
                _cell_text(value).strip() or f"column_{index}"
                for index, value in enumerate(header_values, start=1)
            )
            for row_number, row_values in enumerate(rows, start=2):
                values = tuple(_cell_text(value) for value in row_values)
                if not any(values):
                    continue
                fields = tuple(
                    f"{name}: {value}"
                    for name, value in zip(header, values, strict=False)
                )
                if len(values) > len(header):
                    fields += tuple(
                        f"column_{index}: {value}"
                        for index, value in enumerate(
                            values[len(header) :], start=len(header) + 1
                        )
                    )
                builder.add_record(
                    "\n".join(fields),
                    kind="tabular_row",
                    locations=(
                        SourceLocation(
                            source=_source_path(document),
                            sheet=worksheet.title,
                            row_start=row_number,
                            row_end=row_number,
                            column_start=1,
                            column_end=max(1, len(values)),
                        ),
                    ),
                    attrs={
                        "header": header,
                        "values": values,
                        "sheet": worksheet.title,
                    },
                )
        workbook.close()
        return builder.build()


__all__ = ["XlsxParser"]
